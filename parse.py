from xml.dom import minidom
from openpyxl import Workbook
from openpyxl.styles import Alignment

wb, worksheets = Workbook(), list()

def get_table(name, tables=minidom.parse('eq7.XML').getElementsByTagName('SELECTCASE')):
    for table in tables:
        if table.parentNode.attributes.get('name').value == name:
            return table

prd_all_data = {
    'eq7': {
        'prd_tables': ['FS_TR_EQ7_V6_CAT_PRD_LT_I', 'FS_TR_EQ7_V6_CAT_PRD_LT_M', 'FS_TR_EQ7_V6_CAT_PRD_LT_R', 'FS_TR_EQ7_V6_CAT_PRD_LT_C', 'FS_TR_EQ7_V6_CAT_PRD_LT_O', 'FS_IQ_EQ7_V6_CAT_PRD_LT'],
        'col_headers': ['SEGMENT', '*AND*', 'ACCOUNT TYPE', '*AND*', 'LOAN TYPE\n**OR**', '*AND*', 'KOB\n**OR**'],
        'column_map': [-3, -3, -3, -3, -3 -3],
        'segments': ['Tradeline', 'Tradeline', 'Tradeline', 'Tradeline', 'Tradeline', 'Inquiry'],
        'acct_types': ['Installment', 'Mortgage', 'Revolving', 'Checking', 'Open Ended', 'Irrelevant'] 
    },
    'eqC': {
        'prd_tables': ['FS_TR_EQC_V6_CAT_PRD_NC_I', 'FS_TR_EQC_V6_CAT_PRD_NC_R', 'FS_IQ_EQC_V6_CAT_PRD_KOB'],
        'col_headers': ['SEGMENT', '*AND*', 'ACCOUNT TYPE', '*AND*', 'COMMENT CODES ( CONTAIN )\n**OR**', '*AND*', 'KOB\n**OR**'],
        'column_map': [-3, -3, -1],
        'segments': ['Tradeline', 'Tradeline', 'Inquiry'],
        'acct_types': ['Installment', 'Revolving', 'Irrelevant']
    }
}


# prd_tables = ['FS_TR_EQ7_V6_CAT_PRD_LT_I', 'FS_TR_EQ7_V6_CAT_PRD_LT_M', 'FS_TR_EQ7_V6_CAT_PRD_LT_R', 'FS_TR_EQ7_V6_CAT_PRD_LT_C', 'FS_TR_EQ7_V6_CAT_PRD_LT_O', 'FS_IQ_EQ7_V6_CAT_PRD_LT']
prd_tables = ['FS_TR_EQC_V6_CAT_PRD_NC_I', 'FS_TR_EQC_V6_CAT_PRD_NC_R', 'FS_IQ_EQC_V6_CAT_PRD_KOB']
column_map = [-3, -3, -1]
prd_cats = ['AUT', 'CCB', 'CCO', 'CCR', 'CCS', 'EDU', 'LOC', 'MAR', 'MHL', 'MLC', 'MRT', 'OIN', 'ORE', 'ORV']
acct_types = ['Installment', 'Revolving', 'Irrelevant']
# acct_types = ['Installment', 'Mortgage', 'Revolving', 'Checking', 'Open Ended', 'Irrelevant']
# segments = ['Tradeline', 'Tradeline', 'Tradeline', 'Tradeline', 'Tradeline', 'Inquiry']
segments = ['Tradeline', 'Tradeline', 'Inquiry']
parsed_xml = minidom.parse('eq7.XML')

worksheets.append(wb.create_sheet('PRD -->'))
for cat_i, cat in enumerate(prd_cats):
    worksheets.append(wb.create_sheet(cat))
    worksheets[cat_i+1].append(['SEGMENT', '*AND*', 'ACCOUNT TYPE', '*AND*', 'COMMENT CODES ( CONTAIN )\n**OR**', '*AND*', 'KOB\n**OR**'])
    for table_i in range(len(prd_tables)):
        target_table, target_cat, cond_dict = prd_tables[table_i], cat, dict()
        tables = parsed_xml.getElementsByTagName('SELECTCASE')
        direct_returns = {}
        indirect_returns = {}

        def insert_cond(cond, cond_dict):
            split_cond = cond.split(' ')
            param_op, target = ' '.join(split_cond[:2]), split_cond[2]
            cond_dict[param_op] = cond_dict.get(param_op, []) + [target]

        for table in tables:
            cases = table.getElementsByTagName('CASE')
            if table.parentNode.attributes.get('name').value == target_table:
                if target_table == "FS_TR_EQ5_V6_CAT_PRD_NC_R":
                    import pdb;pdb.set_trace()
                for case in cases:
                    get = case.attributes.get
                    if get('result').value.strip()[1:-1] == target_cat:
                        insert_cond(get('test').value, direct_returns)
                    elif get('returntype').value == 'Attribute':
                        ret_val = get('result').value
                        indirect_returns[ret_val] = indirect_returns.get(ret_val, []) + [get('test').value]

        paired_conds = dict()
        for return_table, conditions in indirect_returns.items():
            paired_conds[return_table] = conditions + ['&&']
            is_table = False
            for table in tables:
                cases = table.getElementsByTagName('CASE')
                if table.parentNode.attributes.get('name').value == return_table:
                    is_table = True
                    found_target_cat = False
                    for case in cases:
                        if case.attributes.get('result').value.strip()[1:-1] == target_cat:
                            found_target_cat = True
                            paired_conds[return_table].append(case.attributes.get('test').value)
                        elif case.attributes.get('returntype').value == 'Attribute':
                            import pdb;pdb.set_trace()
                    if not found_target_cat:
                        del paired_conds[return_table]
                    break
            if not is_table:
                paired_conds[return_table][0] += '*********'

        print(prd_tables[table_i])
        for param_op, targets in direct_returns.items():
            cleaned_targets = list()
            for tar_i, tar in enumerate(targets):
                tar = f'{tar}    '
                if (tar_i+1) % 15 == 0:
                    tar += '\n'
                cleaned_targets.append(tar.replace("'", '').replace('"', ''))
            targets_str = ''.join(cleaned_targets)
            new_ws_row = [segments[table_i], '', acct_types[table_i], '', 'Irrelevant', '', 'Irrelevant']
            new_ws_row[column_map[table_i]] = targets_str
            worksheets[cat_i+1].append(new_ws_row)
            print(f"{param_op}  :  {targets_str}")
        print('\n')
            

        transformed_indirects = dict()
        for table, conditions in paired_conds.items():
            transformed_indirects[table] = [dict(), dict()]
            slice_index = conditions.index('&&')
            tiered_conds = [conditions[:slice_index], conditions[slice_index+1:]]
            for i, conds in enumerate(tiered_conds):
                [insert_cond(cond, transformed_indirects[table][i]) for cond in conds]

        for table in transformed_indirects:   
            conds = list(transformed_indirects[table][0].items()) + list(transformed_indirects[table][1].items())
            if len(conds) != 2:
                import pdb;pdb.set_trace()
            else:
                cleaned_targets = [list(), list()]    
                for col_i, (param_op, targets) in enumerate(conds):
                    for tar_i, tar in enumerate(targets):
                        tar = f'{tar}    '
                        if (tar_i+1) % 15 == 0:
                            tar += '\n'
                        cleaned_targets[col_i].append(tar.replace("'", '').replace('"', ''))
                    cleaned_targets[col_i] = ''.join(cleaned_targets[col_i])     
                    targets_str = '    '.join(targets).replace("'", '').replace('"', '')
                    print(f"{param_op}  :  {targets_str}")
                new_ws_row = [segments[table_i], '', acct_types[table_i], '', cleaned_targets[0], '', cleaned_targets[1]]
                worksheets[cat_i+1].append(new_ws_row)
            print('\n')
        print('\n\n')



master = [
    {
        'SPC': 
        {   
            'Categories': ['ARF', 'AUL', 'COM', 'CON', 'SDE', 'FOB', 'GVT', 'MED', 'MMH', 'OIL', 'OLS', 'RCV'],
            'Tables By Segment':    
            [
                {
                    'Segment': 'Tradeline',
                    'Columns': 'LOAN TYPE & COMMENT CODE (CONTAINS)',
                    'Tables By Account Type': 
                    [
                        {
                            'acct_type': 'Installment',             
                            'tables': (get_table('FS_TR_EQ7_V6_CAT_SPC_LT_I'), get_table('FS_TR_EQ7_V6_CAT_SPC_NC_I'))
                        },
                        {
                            'acct_type': 'Revolving', 
                            'tables': (get_table('FS_TR_EQ7_V6_CAT_SPC_LT_R'), get_table('FS_TR_EQ7_V6_CAT_SPC_NC_R'))
                        }
                    ]
                },
                {
                    'Segment': 'Inquiry',
                    'Columns': 'LOAN TYPE & KOB',
                    'Tables By Account Type': 
                    [
                        {
                            'acct_type': 'Irrelevant', 
                            'tables': (get_table('FS_IQ_EQ7_V6_CAT_SPC_LT'), get_table('FS_IQ_EQ7_V6_CAT_SPC_KOB'))
                        }
                    ]
                }
            ]
        }
    },
    {
        'SRC':
        {   
            'Categories': ['BNK', 'CRU', 'NBF', 'PFN', 'TOT', 'RTL', 'TCL', 'TCM', 'UTL'],
            'Tables By Segment':    
            [
                {
                    'Segment': 'Tradeline',
                    'Columns': 'CREDCLASS & KOB',
                    'Tables By Account Type':
                    [
                        {
                            'acct_type': 'Installment',             
                            'tables': (get_table('FS_TR_EQ7_V6_CAT_SRC_CCC_I'), get_table('FS_TR_EQ7_V6_CAT_SRC_KOB_I'))
                        },
                        {
                            'acct_type': 'Revolving', 
                            'tables': (get_table('FS_TR_EQ7_V6_CAT_SRC_CCC_R'), get_table('FS_TR_EQ7_V6_CAT_SRC_KOB_R'))
                        }
                    ]
                },
                {
                    'Segment': 'Inquiry',
                    'Columns': 'KOB',
                    'Specified Indices': {'Irrelevant': -3, 'Codes': -1},
                    'Tables By Account Type': 
                    [
                        {
                            'acct_type': 'Irrelevant', 
                            'tables': (get_table('FS_IQ_EQ7_V6_CAT_SRC'),)
                        }
                    ]
                }
            ]
        }
    }
]

for data in master:
    cat_title = list(data.keys())[0]
    worksheets.append(wb.create_sheet(f'{cat_title} -->'))
    cats, segments = data[cat_title]['Categories'], data[cat_title]['Tables By Segment']
    for cat in cats:
        worksheets.append(wb.create_sheet(cat))
        init_cols = list() 
        base_cols = ['SEGMENT', '*AND*', 'ACCOUNT TYPE', '*AND*']
        for segment_data in segments:
            possibly_new_cols = [col.strip() for col in segment_data['Columns'].split('&')]
            if init_cols == list():
                worksheets[-1].append(base_cols + [possibly_new_cols[0], '*AND*', possibly_new_cols[1]])
            if init_cols != possibly_new_cols and init_cols != list() and len(possibly_new_cols) > 1:
                worksheets[-1].append(['', '', '', '', possibly_new_cols[0], '*AND*', possibly_new_cols[1]])
            init_cols = possibly_new_cols
            for account_data in segment_data['Tables By Account Type']:
                codes = [list(), list()]
                for i, table in enumerate(account_data['tables']):
                    for case in table.getElementsByTagName('CASE'):
                        case_return = case.attributes.get('result').value.strip()[1:-1]
                        if case_return == cat:
                            codes[i].append(case.attributes.get('value').value.replace('"', '').replace("'", ''))
                for i, code_set in enumerate(codes):
                    for j, code in enumerate(code_set):
                        code_set[j] = f'{code}    '
                        if (j+1) % 10 == 0:
                            code_set[j] += '\n\n'
                    codes[i] = ''.join(code_set)
                first_row = [segment_data['Segment'], '', account_data['acct_type'], '', codes[0], '', 'Irrelevant']
                second_row = [segment_data['Segment'], '', account_data['acct_type'], '', 'Unrecognized', '', codes[1]]
                undefined_row = [segment_data['Segment'], '', account_data['acct_type'], '', '', '', '']
                if 'Specified Indices' not in segment_data:
                    if len(codes[0]):
                        worksheets[-1].append(first_row)
                    if len(codes[1]):
                        worksheets[-1].append(second_row)
                    if not len(codes[0]) and not len(codes[1]): 
                        worksheets[-1].append(undefined_row)
                else:
                    undefined_row[segment_data['Specified Indices']['Irrelevant']] = 'Irrelevant'
                    undefined_row[segment_data['Specified Indices']['Codes']] = codes[0]
                    worksheets[-1].append(undefined_row)
                    

                            
wb.save('eq7_cat.xlsx')

